"""Discover, verify, and ingest New York Fed HHDC workbook histories."""

from __future__ import annotations

import argparse
import json
import re
from collections.abc import Mapping, Sequence
from dataclasses import asdict, dataclass
from datetime import UTC, date, datetime
from io import BytesIO
from math import isfinite
from pathlib import Path
from typing import Any
from urllib.parse import urljoin
from zipfile import BadZipFile

import polars as pl
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from dfri.ingest.http import HttpTransport
from dfri.ingest.registry import NyFedSeriesDefinition, load_nyfed_series
from dfri.lake.store import AppendOnlyParquetStore

NYFED_HHDC_PAGE = "https://www.newyorkfed.org/householdcredit/hhdc-iframe"
NYFED_BASE = "https://www.newyorkfed.org"
NYFED_WORKBOOK_RE = re.compile(
    r"/medialibrary/interactives/householdcredit/data/xls/"
    r"HHD_C_Report_(\d{4}Q[1-4])(?:\.xlsx)?",
    flags=re.IGNORECASE,
)
NYFED_ATTRIBUTION = "New York Fed Consumer Credit Panel/Equifax"
DEFAULT_START = date(2015, 1, 1)


class NyFedContractError(RuntimeError):
    """The HHDC discovery page, workbook, or stored snapshot violated its contract."""


@dataclass(frozen=True)
class NyFedWorkbookData:
    report_period: date
    source_url: str
    checksum: str
    retrieved_at: str
    content: bytes


@dataclass(frozen=True)
class NyFedHistoryReceipt:
    source_url: str
    checksum: str
    report_period: date
    snapshot_at: datetime
    series_count: int
    row_count: int
    already_present: bool
    batch_path: Path | None
    content_hash: str | None


@dataclass(frozen=True)
class NyFedHistoryValidation:
    total_rows: int
    snapshot_batches: int
    rows_by_series: dict[str, int]
    latest_period_by_series: dict[str, date]


class NyFedClient:
    def __init__(self, transport: HttpTransport) -> None:
        self._transport = transport

    def fetch(self) -> NyFedWorkbookData:
        page = self._transport.get(
            NYFED_HHDC_PAGE,
            headers={"Accept": "text/html"},
        )
        workbook_url = discover_workbook_url(page.content)
        workbook = self._transport.get(
            workbook_url,
            headers={"Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        )
        return NyFedWorkbookData(
            report_period=report_period_from_url(workbook.source_url),
            source_url=workbook.source_url,
            checksum=workbook.checksum,
            retrieved_at=workbook.retrieved_at.isoformat(),
            content=workbook.content,
        )


class NyFedHistoryIngestor:
    def __init__(
        self,
        store: AppendOnlyParquetStore,
        definitions: Sequence[NyFedSeriesDefinition] | None = None,
    ) -> None:
        self._store = store
        self._definitions = tuple(definitions or load_nyfed_series())
        self._known_rows: dict[tuple[str, str], pl.DataFrame] = {}
        self._known_batches: dict[tuple[str, str], Path | None] = {}
        self._load_existing_index()

    def fetch(self, client: NyFedClient, *, start: date = DEFAULT_START) -> NyFedHistoryReceipt:
        return self.ingest(client.fetch(), start=start)

    def ingest(
        self, data: NyFedWorkbookData, *, start: date = DEFAULT_START
    ) -> NyFedHistoryReceipt:
        rows = nyfed_history_rows(data, definitions=self._definitions, start=start)
        identity = (data.source_url, data.checksum)
        matching = self._known_rows.get(identity)
        if matching is not None:
            _verify_existing_rows(matching, rows)
            batch_path = self._known_batches.get(identity)
            snapshot_at = matching["ingested_at"].min()
            if not isinstance(snapshot_at, datetime):
                raise NyFedContractError("Stored NY Fed snapshot timestamp is invalid")
            return NyFedHistoryReceipt(
                source_url=data.source_url,
                checksum=data.checksum,
                report_period=data.report_period,
                snapshot_at=snapshot_at,
                series_count=matching["series_id"].n_unique(),
                row_count=len(rows),
                already_present=True,
                batch_path=batch_path,
                content_hash=(
                    batch_path.stem.removeprefix("batch-") if batch_path is not None else None
                ),
            )

        write = self._store.append("raw_observations", rows)
        stored = pl.read_parquet(write.path)
        self._known_rows[identity] = stored
        self._known_batches[identity] = write.path
        snapshot_at = stored["ingested_at"].min()
        if not isinstance(snapshot_at, datetime):
            raise NyFedContractError("Stored NY Fed snapshot timestamp is invalid")
        return NyFedHistoryReceipt(
            source_url=data.source_url,
            checksum=data.checksum,
            report_period=data.report_period,
            snapshot_at=snapshot_at,
            series_count=stored["series_id"].n_unique(),
            row_count=write.row_count,
            already_present=write.already_present,
            batch_path=write.path,
            content_hash=write.content_hash,
        )

    def _load_existing_index(self) -> None:
        paths = sorted((self._store.root / "raw_observations").glob("batch-*.parquet"))
        for path in paths:
            frame = pl.read_parquet(path).filter(pl.col("source") == "NYFED")
            if frame.is_empty():
                continue
            for source_url, checksum in (
                frame.select(["source_url", "checksum"]).unique().iter_rows()
            ):
                identity = (str(source_url), str(checksum))
                matching = frame.filter(
                    (pl.col("source_url") == source_url) & (pl.col("checksum") == checksum)
                )
                prior = self._known_rows.get(identity)
                if prior is None:
                    self._known_rows[identity] = matching
                    self._known_batches[identity] = path
                else:
                    self._known_rows[identity] = pl.concat([prior, matching], how="vertical")
                    self._known_batches[identity] = None


def discover_workbook_url(content: bytes) -> str:
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise NyFedContractError("NY Fed HHDC discovery page is not UTF-8") from exc
    matches = {match.group(0) for match in NYFED_WORKBOOK_RE.finditer(text)}
    if len(matches) != 1:
        raise NyFedContractError(
            f"NY Fed HHDC discovery found {len(matches)} current workbook endpoints"
        )
    return urljoin(NYFED_BASE, matches.pop())


def report_period_from_url(source_url: str) -> date:
    match = NYFED_WORKBOOK_RE.search(source_url)
    if match is None or not source_url.startswith(NYFED_BASE):
        raise NyFedContractError(f"NY Fed HHDC workbook URL mismatch: {source_url!r}")
    return _report_period(match.group(1))


def nyfed_history_rows(
    data: NyFedWorkbookData,
    *,
    definitions: Sequence[NyFedSeriesDefinition],
    start: date,
) -> list[dict[str, object]]:
    snapshot_at = _snapshot_at(data.retrieved_at)
    _validate_source(data, snapshot_at)
    applicable = tuple(definitions)
    if not applicable:
        raise NyFedContractError("NY Fed HHDC definitions are empty")
    _validate_definitions(applicable, data)

    try:
        workbook = load_workbook(BytesIO(data.content), read_only=False, data_only=True)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise NyFedContractError("NY Fed HHDC workbook is not a valid XLSX file") from exc

    grouped: dict[str, list[NyFedSeriesDefinition]] = {}
    for definition in applicable:
        grouped.setdefault(definition.expected_source_attributes["sheet"], []).append(definition)
    missing = set(grouped) - set(workbook.sheetnames)
    if missing:
        raise NyFedContractError(f"NY Fed HHDC workbook is missing sheets: {sorted(missing)}")
    if not _has_attribution(workbook, tuple(grouped)):
        raise NyFedContractError("NY Fed HHDC workbook attribution is missing")

    rows: list[dict[str, object]] = []
    for sheet_name, sheet_definitions in sorted(grouped.items()):
        sheet = workbook[sheet_name]
        rows.extend(
            _sheet_rows(
                sheet,
                sheet_definitions,
                start=start,
                report_period=data.report_period,
                snapshot_at=snapshot_at,
                source_url=data.source_url,
                checksum=data.checksum,
            )
        )
    return rows


def validate_nyfed_history(
    store: AppendOnlyParquetStore,
    definitions: Sequence[NyFedSeriesDefinition] | None = None,
) -> NyFedHistoryValidation:
    expected_definitions = tuple(definitions or load_nyfed_series())
    expected = {definition.series_id: definition for definition in expected_definitions}
    frame = store.read_table("raw_observations").filter(pl.col("source") == "NYFED")
    if frame.is_empty():
        raise NyFedContractError("NY Fed HHDC history is empty")
    if set(frame["series_id"].unique().to_list()) != set(expected):
        raise NyFedContractError("NY Fed HHDC series coverage is incomplete")
    if frame.select(pl.col("checksum").str.contains(r"^[0-9a-f]{64}$").all()).item() is not True:
        raise NyFedContractError("NY Fed HHDC history contains an invalid checksum")
    if frame.select((pl.col("release_date") == pl.col("ingested_at")).all()).item() is not True:
        raise NyFedContractError("NY Fed HHDC history has a false historical release timestamp")
    if (
        frame.select((pl.col("vintage_date") == pl.col("ingested_at").dt.date()).all()).item()
        is not True
    ):
        raise NyFedContractError("NY Fed HHDC history has an invalid snapshot vintage")
    if any(not isfinite(value) or value < 0 for value in frame["value"].to_list()):
        raise NyFedContractError("NY Fed HHDC history contains an invalid value")
    identities = frame.select(["source_url", "checksum", "series_id", "obs_period"])
    if identities.is_duplicated().any():
        raise NyFedContractError("NY Fed HHDC history contains duplicate source observations")

    for series_key, series in frame.group_by("series_id"):
        series_id = series_key[0] if isinstance(series_key, tuple) else series_key
        definition = expected[str(series_id)]
        if series.select((pl.col("unit") == definition.units).all()).item() is not True:
            raise NyFedContractError(f"NY Fed HHDC unit changed: {series_id}")

    batch_count = 0
    for (source_url, _checksum), snapshot in frame.group_by(["source_url", "checksum"]):
        snapshot_times = snapshot["ingested_at"].unique().to_list()
        if len(snapshot_times) != 1 or not isinstance(snapshot_times[0], datetime):
            raise NyFedContractError("NY Fed source response has multiple snapshot times")
        report_period = report_period_from_url(str(source_url))
        _validate_stored_periods(snapshot, expected, report_period)
        batch_count += 1

    rows_by_series = {
        str(series_id): int(count)
        for series_id, count in frame.group_by("series_id").len().iter_rows()
    }
    latest_by_series = {
        str(series_id): period
        for series_id, period in frame.group_by("series_id")
        .agg(pl.col("obs_period").max())
        .iter_rows()
        if isinstance(period, date)
    }
    if len(latest_by_series) != len(expected):
        raise NyFedContractError("NY Fed latest-period calculation is incomplete")
    return NyFedHistoryValidation(
        total_rows=frame.height,
        snapshot_batches=batch_count,
        rows_by_series=dict(sorted(rows_by_series.items())),
        latest_period_by_series=dict(sorted(latest_by_series.items())),
    )


def _sheet_rows(
    sheet: Any,
    definitions: Sequence[NyFedSeriesDefinition],
    *,
    start: date,
    report_period: date,
    snapshot_at: datetime,
    source_url: str,
    checksum: str,
) -> list[dict[str, object]]:
    first_attrs = definitions[0].expected_source_attributes
    if _cell_text(sheet.cell(1, 1).value) != first_attrs["sheet_title"]:
        raise NyFedContractError(f"NY Fed sheet title changed: {sheet.title}")
    if _cell_text(sheet.cell(2, 1).value) != first_attrs["unit_label"]:
        raise NyFedContractError(f"NY Fed sheet unit label changed: {sheet.title}")
    header_row = int(first_attrs["header_row"])
    headers = {
        _cell_text(sheet.cell(header_row, column).value): column
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(header_row, column).value is not None
    }
    period_sets: list[set[date]] = []
    result: list[dict[str, object]] = []
    for definition in definitions:
        attrs = definition.expected_source_attributes
        if (
            attrs["sheet_title"] != first_attrs["sheet_title"]
            or attrs["unit_label"] != first_attrs["unit_label"]
            or attrs["header_row"] != first_attrs["header_row"]
        ):
            raise NyFedContractError(f"NY Fed sheet definitions disagree: {sheet.title}")
        column = headers.get(attrs["header"])
        if column is None:
            raise NyFedContractError(
                f"NY Fed header changed for {definition.series_id}: {attrs['header']!r}"
            )
        selected: dict[date, float] = {}
        for row_number in range(header_row + 1, sheet.max_row + 1):
            raw_period = sheet.cell(row_number, 1).value
            if raw_period is None:
                continue
            period = _quarter_end(raw_period)
            if period < start:
                continue
            raw_value = sheet.cell(row_number, column).value
            if isinstance(raw_value, bool) or not isinstance(raw_value, int | float):
                raise NyFedContractError(
                    f"NY Fed value is missing or nonnumeric: {definition.series_id} {raw_period}"
                )
            value = float(raw_value)
            if not isfinite(value) or value < 0 or (definition.units == "Percent" and value > 100):
                raise NyFedContractError(f"NY Fed value is invalid: {definition.series_id}")
            if period in selected:
                raise NyFedContractError(f"NY Fed period is duplicated: {definition.series_id}")
            selected[period] = value
        if not selected:
            raise NyFedContractError(
                f"NY Fed history has no rows since {start}: {definition.series_id}"
            )
        _validate_periods(
            selected,
            start,
            report_period,
            int(attrs["max_lag_quarters"]),
            definition.series_id,
        )
        period_sets.append(set(selected))
        result.extend(
            {
                "source": "NYFED",
                "series_id": definition.series_id,
                "obs_period": period,
                "value": value,
                "unit": definition.units,
                "release_date": snapshot_at,
                "vintage_date": snapshot_at.date(),
                "ingested_at": snapshot_at,
                "source_url": source_url,
                "checksum": checksum,
            }
            for period, value in sorted(selected.items())
        )
    if any(periods != period_sets[0] for periods in period_sets[1:]):
        raise NyFedContractError(f"NY Fed series periods disagree within sheet: {sheet.title}")
    return result


def _validate_definitions(
    definitions: Sequence[NyFedSeriesDefinition], data: NyFedWorkbookData
) -> None:
    required = {
        "source_page",
        "workbook_path_pattern",
        "verified_workbook_url",
        "verified_report_period",
        "sheet",
        "sheet_title",
        "unit_label",
        "header_row",
        "header",
        "max_lag_quarters",
    }
    if any(set(definition.expected_source_attributes) != required for definition in definitions):
        raise NyFedContractError("NY Fed series attributes are incomplete")
    if any(definition.frequency != "Quarterly" for definition in definitions):
        raise NyFedContractError("NY Fed series frequency changed")
    verified_periods = {
        _report_period(definition.expected_source_attributes["verified_report_period"])
        for definition in definitions
    }
    if len(verified_periods) != 1 or data.report_period < verified_periods.pop():
        raise NyFedContractError("NY Fed report period predates the verified registry contract")
    if any(
        definition.expected_source_attributes["source_page"] != NYFED_HHDC_PAGE
        or definition.expected_source_attributes["workbook_path_pattern"]
        != "/medialibrary/interactives/householdcredit/data/xls/HHD_C_Report_YYYYQn"
        for definition in definitions
    ):
        raise NyFedContractError("NY Fed source endpoint contract changed")
    if any(
        data.report_period
        == _report_period(definition.expected_source_attributes["verified_report_period"])
        and data.source_url != definition.expected_source_attributes["verified_workbook_url"]
        for definition in definitions
    ):
        raise NyFedContractError("NY Fed verified workbook URL changed")


def _validate_source(data: NyFedWorkbookData, snapshot_at: datetime) -> None:
    if report_period_from_url(data.source_url) != data.report_period:
        raise NyFedContractError("NY Fed URL and report period disagree")
    if re.fullmatch(r"[0-9a-f]{64}", data.checksum) is None:
        raise NyFedContractError("NY Fed checksum is not lowercase SHA-256")
    age_days = (snapshot_at.date() - data.report_period).days
    if not 20 <= age_days <= 190:
        raise NyFedContractError("NY Fed workbook is stale or implausibly current")


def _has_attribution(workbook: Any, sheet_names: Sequence[str]) -> bool:
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            if any(NYFED_ATTRIBUTION in value for value in row if isinstance(value, str)):
                return True
    return False


def _validate_periods(
    selected: Mapping[date, float],
    start: date,
    report_period: date,
    lag_quarters: int,
    series_id: str,
) -> None:
    expected_latest = _shift_quarters(report_period, -lag_quarters)
    if max(selected) != expected_latest:
        raise NyFedContractError(
            "NY Fed series latest period does not match its declared lag: "
            f"{series_id} expected {expected_latest}, found {max(selected)}"
        )
    expected_first = _quarter_containing(start)
    if min(selected) != expected_first:
        raise NyFedContractError(
            "NY Fed series does not begin at the requested start quarter: "
            f"{series_id} expected {expected_first}, found {min(selected)}"
        )
    if set(selected) != set(_quarter_sequence(expected_first, expected_latest)):
        raise NyFedContractError(f"NY Fed series has a gap in quarterly coverage: {series_id}")


def _validate_stored_periods(
    snapshot: pl.DataFrame,
    definitions: Mapping[str, NyFedSeriesDefinition],
    report_period: date,
) -> None:
    for series_key, series in snapshot.group_by("series_id"):
        series_id = series_key[0] if isinstance(series_key, tuple) else series_key
        definition = definitions[str(series_id)]
        periods = set(series["obs_period"].to_list())
        latest = _shift_quarters(
            report_period,
            -int(definition.expected_source_attributes["max_lag_quarters"]),
        )
        if max(periods) != latest or set(_quarter_sequence(min(periods), latest)) != periods:
            raise NyFedContractError(f"Stored NY Fed period coverage is invalid: {series_id}")


def _snapshot_at(raw: str) -> datetime:
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError as exc:
        raise NyFedContractError("NY Fed retrieval timestamp is invalid") from exc
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        raise NyFedContractError("NY Fed retrieval timestamp must be timezone-aware")
    return parsed.astimezone(UTC)


def _report_period(raw: str) -> date:
    match = re.fullmatch(r"(\d{4})Q([1-4])", raw)
    if match is None:
        raise NyFedContractError(f"Invalid NY Fed report period: {raw!r}")
    return _quarter_end_from_parts(int(match.group(1)), int(match.group(2)))


def _quarter_end(raw: object) -> date:
    if not isinstance(raw, str):
        raise NyFedContractError(f"Invalid NY Fed quarterly period: {raw!r}")
    match = re.fullmatch(r"(\d{2}):Q([1-4])", raw.strip())
    if match is None:
        raise NyFedContractError(f"Invalid NY Fed quarterly period: {raw!r}")
    return _quarter_end_from_parts(2000 + int(match.group(1)), int(match.group(2)))


def _quarter_end_from_parts(year: int, quarter: int) -> date:
    month = quarter * 3
    day = 31 if month in {3, 12} else 30
    return date(year, month, day)


def _quarter_containing(value: date) -> date:
    quarter = (value.month - 1) // 3 + 1
    return _quarter_end_from_parts(value.year, quarter)


def _shift_quarters(value: date, offset: int) -> date:
    quarter_index = value.year * 4 + (value.month - 1) // 3 + offset
    year, quarter_zero = divmod(quarter_index, 4)
    return _quarter_end_from_parts(year, quarter_zero + 1)


def _quarter_sequence(first: date, last: date) -> tuple[date, ...]:
    periods: list[date] = []
    current = first
    while current <= last:
        periods.append(current)
        current = _shift_quarters(current, 1)
    return tuple(periods)


def _cell_text(value: object) -> str:
    return value.strip() if isinstance(value, str) else ""


def _verify_existing_rows(
    existing: pl.DataFrame, expected_rows: Sequence[Mapping[str, object]]
) -> None:
    columns = ["source", "series_id", "obs_period", "value", "unit", "source_url", "checksum"]
    expected = {tuple(row[column] for column in columns) for row in expected_rows}
    actual = {tuple(row) for row in existing.select(columns).iter_rows()}
    if len(existing) != len(expected_rows) or actual != expected:
        raise NyFedContractError(
            "Stored NY Fed snapshot with the same source checksum does not match the live parse"
        )


def _parse_date(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("expected YYYY-MM-DD") from exc


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--start", type=_parse_date, default=DEFAULT_START)
    parser.add_argument("--lake-root", type=Path, default=Path(".local/lake/raw"))
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    store = AppendOnlyParquetStore(args.lake_root)
    with HttpTransport(min_interval_seconds=0.5) as transport:
        receipt = NyFedHistoryIngestor(store).fetch(NyFedClient(transport), start=args.start)
        validation = validate_nyfed_history(store)
    print(
        json.dumps(
            {"snapshot": asdict(receipt), "validation": asdict(validation)},
            sort_keys=True,
            default=str,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
